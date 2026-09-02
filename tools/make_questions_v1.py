#!/usr/bin/env python3
"""2026-09-02 ウェアラブルカメラ素材(raw/)から作った第1版の問題(本試験20+デモ3)を tools/questions.xlsx に書く。
言語: ja + en + vi + id（es は空=日本語表示）。2026-09-02 社長判断: 全問を試験に採用(剖検・採血含む)、q18事務所での吸い上げ/q26洗い場は正しい、同一動画への複数問(q27〜q30)は許容 → placeholder=0。
既存の questions.xlsx は --force で上書き。以後の修正は Excel を直接編集して build_from_xlsx.py。
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_from_xlsx import HEAD, XLSX  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, ValueError):
    pass

if XLSX.exists() and "--force" not in sys.argv:
    raise SystemExit(f"{XLSX} は既にあります(--force で上書き)")

# 各問: (id, category, answer(1-4), explanation_ja, {lang: (question, [opt1..opt4])})
# 選択肢の並びは全言語で同じ（順序を崩すと採点事故）。正解は answer 番号。
Q = [
    ("q01", "衛生・防疫", 1,
     "豚舎の入口で、外で使う長靴と豚舎の中で使う長靴を分けています。外の汚れ（病気のもと）を豚舎に持ち込まないためです。",
     {"ja": ("動画の人は、豚舎に入る前に何をしていますか？", ["長靴をはきかえている", "手を洗っている", "豚にえさをやっている", "服を着がえている"]),
      "en": ("What is the person in the video doing before entering the pig house?", ["Changing boots", "Washing hands", "Feeding the pigs", "Changing clothes"]),
      "vi": ("Người trong video làm gì trước khi vào chuồng heo?", ["Thay ủng", "Rửa tay", "Cho heo ăn", "Thay quần áo"]),
      "id": ("Apa yang dilakukan orang dalam video sebelum masuk kandang babi?", ["Mengganti sepatu bot", "Mencuci tangan", "Memberi makan babi", "Mengganti pakaian"])}),
    ("q02", "衛生・防疫", 1,
     "注射などの作業の前に使い捨て手袋を着けます。手の汚れを薬や豚につけないため、また薬が手につかないためです。",
     {"ja": ("動画の人は、作業の前に何を着けていますか？", ["使い捨て手袋", "マスク", "ヘルメット", "エプロン"]),
      "en": ("What is the person in the video putting on before the work?", ["Disposable gloves", "A mask", "A helmet", "An apron"]),
      "vi": ("Người trong video đeo gì trước khi làm việc?", ["Găng tay dùng một lần", "Khẩu trang", "Mũ bảo hộ", "Tạp dề"]),
      "id": ("Apa yang dipakai orang dalam video sebelum bekerja?", ["Sarung tangan sekali pakai", "Masker", "Helm", "Celemek"])}),
    ("q03", "衛生・防疫", 1,
     "注射の前に注射器と針を組み立てて準備をしています。針がゆるんでいないか確認します。",
     {"ja": ("動画では何をしていますか？", ["注射器を組み立てている", "注射器を洗っている", "豚に注射している", "薬を捨てている"]),
      "en": ("What is being done in the video?", ["Assembling a syringe", "Washing a syringe", "Injecting a pig", "Throwing away medicine"]),
      "vi": ("Trong video đang làm gì?", ["Lắp ráp ống tiêm", "Rửa ống tiêm", "Tiêm cho heo", "Vứt bỏ thuốc"]),
      "id": ("Apa yang sedang dilakukan dalam video?", ["Merakit alat suntik", "Mencuci alat suntik", "Menyuntik babi", "Membuang obat"])}),
    ("q04", "衛生・防疫", 1,
     "薬のびんに針をさして、注射器に薬を吸い上げています。決められた量を正確に吸います。",
     {"ja": ("動画では何をしていますか？", ["薬のびんから注射器に薬を吸っている", "薬を豚に飲ませている", "注射器を消毒している", "びんに水を入れている"]),
      "en": ("What is being done in the video?", ["Drawing medicine from the bottle into the syringe", "Giving medicine to a pig by mouth", "Disinfecting the syringe", "Filling the bottle with water"]),
      "vi": ("Trong video đang làm gì?", ["Hút thuốc từ lọ vào ống tiêm", "Cho heo uống thuốc", "Khử trùng ống tiêm", "Đổ nước vào lọ"]),
      "id": ("Apa yang sedang dilakukan dalam video?", ["Menyedot obat dari botol ke alat suntik", "Memberi obat ke babi lewat mulut", "Mendisinfeksi alat suntik", "Mengisi botol dengan air"])}),
    ("q05", "子豚・育成", 1,
     "子豚に注射をしたあと、赤いスプレーで印をつけています。印は「注射がすんだ」しるしです。",
     {"ja": ("動画では、子豚に何をしていますか？", ["注射をして、赤いスプレーで印をつけている", "耳に番号をつけている", "体重をはかっている", "えさをやっている"]),
      "en": ("What is being done to the piglets in the video?", ["Injecting them and marking them with red spray", "Putting number tags on their ears", "Weighing them", "Feeding them"]),
      "vi": ("Trong video đang làm gì với heo con?", ["Tiêm rồi đánh dấu bằng sơn xịt màu đỏ", "Gắn số vào tai", "Cân heo", "Cho heo ăn"]),
      "id": ("Apa yang dilakukan pada anak babi dalam video?", ["Menyuntik lalu menandai dengan semprotan merah", "Memasang nomor di telinga", "Menimbang berat", "Memberi makan"])}),
    ("q06", "子豚・育成", 1,
     "印がないと、同じ豚に2回注射したり、注射をしていない豚を見のがしたりします。印を見て「まだの豚」だけに注射します。",
     {"ja": ("注射のあと、子豚に赤いスプレーをかけるのはなぜですか？", ["注射がすんだ豚を見分けるため", "豚を元気にするため", "虫をよけるため", "きずを消毒するため"]),
      "en": ("Why is red spray put on the piglet after the injection?", ["To tell which pigs have already been injected", "To make the pig healthier", "To keep insects away", "To disinfect a wound"]),
      "vi": ("Vì sao xịt sơn đỏ lên heo con sau khi tiêm?", ["Để phân biệt heo đã tiêm", "Để heo khỏe hơn", "Để đuổi côn trùng", "Để sát trùng vết thương"]),
      "id": ("Mengapa anak babi disemprot merah setelah disuntik?", ["Untuk membedakan babi yang sudah disuntik", "Untuk membuat babi lebih sehat", "Untuk mengusir serangga", "Untuk mendisinfeksi luka"])}),
    ("q07", "子豚・育成", 1,
     "赤い印がついている豚は注射がすんでいます。印のない豚をさがして注射します。",
     {"ja": ("動画の豚房で、まだ注射をしていない子豚はどれですか？", ["赤い印がない子豚", "赤い印がある子豚", "寝ている子豚", "小さい子豚"]),
      "en": ("In the pen shown, which piglets have NOT been injected yet?", ["Piglets without a red mark", "Piglets with a red mark", "Piglets that are sleeping", "The smaller piglets"]),
      "vi": ("Trong ô chuồng ở video, heo con nào CHƯA được tiêm?", ["Heo con không có dấu đỏ", "Heo con có dấu đỏ", "Heo con đang ngủ", "Heo con nhỏ"]),
      "id": ("Di kandang dalam video, anak babi mana yang BELUM disuntik?", ["Anak babi tanpa tanda merah", "Anak babi dengan tanda merah", "Anak babi yang tidur", "Anak babi yang kecil"])}),
    ("q08", "繁殖・分娩", 1,
     "ストール（1頭ずつの柵）に入った母豚のとなりで、小さいびん（バイアル）から注射器に薬を吸っています。",
     {"ja": ("動画の場所と作業はどれですか？", ["母豚の豚舎で、注射の薬を吸っている", "子豚の豚舎で、えさを配っている", "分娩房で、子豚を運んでいる", "事務所で、記録を書いている"]),
      "en": ("Where is the video taken, and what is being done?", ["In the sow house, drawing medicine into a syringe", "In the piglet house, giving feed", "In a farrowing pen, carrying piglets", "In the office, writing records"]),
      "vi": ("Video quay ở đâu và đang làm gì?", ["Ở chuồng heo nái, hút thuốc vào ống tiêm", "Ở chuồng heo con, phát thức ăn", "Ở ô đẻ, bế heo con", "Ở văn phòng, ghi sổ"]),
      "id": ("Di mana video ini diambil dan apa yang dilakukan?", ["Di kandang induk, menyedot obat ke alat suntik", "Di kandang anak babi, membagi pakan", "Di kandang beranak, membawa anak babi", "Di kantor, menulis catatan"])}),
    ("q09", "子豚・育成", 1,
     "子豚が動くと針が折れたり、ちがう場所に刺さったりします。片手でしっかりおさえて（保定して）から注射します。",
     {"ja": ("動画では、子豚をどうやって注射していますか？", ["片手で子豚をおさえて、もう片方の手で注射している", "子豚を走らせながら注射している", "二人で子豚を持ち上げている", "子豚をひもでしばっている"]),
      "en": ("How is the piglet being injected in the video?", ["Holding the piglet with one hand and injecting with the other", "Injecting while the piglet runs", "Two people lifting the piglet", "Tying the piglet with a rope"]),
      "vi": ("Trong video, heo con được tiêm như thế nào?", ["Một tay giữ heo con, tay kia tiêm", "Tiêm khi heo con đang chạy", "Hai người nhấc heo con lên", "Buộc heo con bằng dây"]),
      "id": ("Bagaimana anak babi disuntik dalam video?", ["Satu tangan memegang anak babi, tangan lain menyuntik", "Menyuntik sambil anak babi berlari", "Dua orang mengangkat anak babi", "Mengikat anak babi dengan tali"])}),
    ("q10", "衛生・防疫", 1,
     "使った注射器はそのままにせず、分解して洗い、きれいにしてから片づけます。よごれた注射器は病気を広げます。",
     {"ja": ("動画では何をしていますか？", ["使ったあとの注射器を分解して洗っている", "新しい注射器を箱から出している", "注射器に薬を入れている", "注射器を豚に刺している"]),
      "en": ("What is being done in the video?", ["Taking apart and washing a used syringe", "Taking a new syringe out of the box", "Filling a syringe with medicine", "Injecting a pig with the syringe"]),
      "vi": ("Trong video đang làm gì?", ["Tháo rời và rửa ống tiêm đã dùng", "Lấy ống tiêm mới ra khỏi hộp", "Bơm thuốc vào ống tiêm", "Tiêm cho heo"]),
      "id": ("Apa yang sedang dilakukan dalam video?", ["Membongkar dan mencuci alat suntik bekas", "Mengeluarkan alat suntik baru dari kotak", "Mengisi obat ke alat suntik", "Menyuntik babi"])}),
    ("q11", "衛生・防疫", 1,
     "外から来た人（獣医など）は、白い使い捨ての防護服（つなぎ）と長靴を着けて豚舎に入ります。外の病気を持ち込まず、持ち出さないためです。",
     {"ja": ("動画の人が着ているものは何ですか？", ["使い捨ての白い防護服と長靴", "ふだんの服と運動靴", "雨がっぱとサンダル", "作業ズボンだけ"]),
      "en": ("What is the person in the video wearing?", ["A disposable white coverall and boots", "Everyday clothes and sneakers", "A raincoat and sandals", "Only work trousers"]),
      "vi": ("Người trong video mặc gì?", ["Bộ đồ bảo hộ trắng dùng một lần và ủng", "Quần áo thường và giày thể thao", "Áo mưa và dép", "Chỉ quần lao động"]),
      "id": ("Apa yang dipakai orang dalam video?", ["Baju pelindung putih sekali pakai dan sepatu bot", "Pakaian biasa dan sepatu olahraga", "Jas hujan dan sandal", "Hanya celana kerja"])}),
    ("q12", "肉豚・出荷・肉", 1,
     "鼻にワイヤーをかける道具（鼻保定器・スネア）です。大きい豚は手ではおさえられないので、採血や注射のときに使います。",
     {"ja": ("動画で豚の鼻にかけている道具は何のためですか？", ["豚が動かないようにおさえる（保定する）ため", "豚に水を飲ませるため", "豚の体重をはかるため", "豚を運ぶため"]),
      "en": ("What is the tool on the pig's snout in the video for?", ["To hold the pig still (restraint)", "To give the pig water", "To weigh the pig", "To carry the pig"]),
      "vi": ("Dụng cụ móc vào mũi heo trong video dùng để làm gì?", ["Để giữ heo đứng yên (cố định)", "Để cho heo uống nước", "Để cân heo", "Để khiêng heo"]),
      "id": ("Untuk apa alat yang dipasang di hidung babi dalam video?", ["Untuk menahan babi agar tidak bergerak", "Untuk memberi babi minum", "Untuk menimbang babi", "Untuk mengangkut babi"])}),
    ("q13", "衛生・防疫", 1,
     "鼻保定をした豚の首から注射器で血をとっています。血は病気の検査に使います。",
     {"ja": ("動画では、豚に何をしていますか？", ["首から血をとっている（採血）", "首に薬を注射している", "首の温度をはかっている", "首のきずを消毒している"]),
      "en": ("What is being done to the pig in the video?", ["Taking blood from the neck (blood sampling)", "Injecting medicine into the neck", "Measuring the temperature of the neck", "Disinfecting a wound on the neck"]),
      "vi": ("Trong video đang làm gì với con heo?", ["Lấy máu ở cổ (lấy mẫu máu)", "Tiêm thuốc vào cổ", "Đo nhiệt độ ở cổ", "Sát trùng vết thương ở cổ"]),
      "id": ("Apa yang dilakukan pada babi dalam video?", ["Mengambil darah dari leher (pengambilan darah)", "Menyuntik obat ke leher", "Mengukur suhu leher", "Mendisinfeksi luka di leher"])}),
    ("q14", "衛生・防疫", 1,
     "死んだ子豚をならべて、獣医が体を開いて死んだ原因を調べます（剖検）。手袋と防護服をつけて行います。",
     {"ja": ("動画で、床にならんでいるのは何ですか？このあと何をしますか？", ["死んだ子豚。原因を調べるために体を開く（解剖）", "寝ている子豚。起こしてえさをやる", "生まれたばかりの子豚。母豚に返す", "病気の子豚。注射をする"]),
      "en": ("What is lying on the floor in the video, and what happens next?", ["Dead piglets. They will be opened to find the cause of death (necropsy)", "Sleeping piglets. They will be woken and fed", "Newborn piglets. They will be returned to the sow", "Sick piglets. They will be injected"]),
      "vi": ("Trong video, thứ nằm trên sàn là gì và tiếp theo làm gì?", ["Heo con đã chết. Sẽ mổ để tìm nguyên nhân chết (mổ khám)", "Heo con đang ngủ. Sẽ đánh thức và cho ăn", "Heo con mới sinh. Sẽ trả về heo mẹ", "Heo con bị bệnh. Sẽ tiêm thuốc"]),
      "id": ("Apa yang terbaring di lantai dalam video, dan apa yang dilakukan selanjutnya?", ["Anak babi mati. Akan dibedah untuk mencari penyebab kematian (nekropsi)", "Anak babi tidur. Akan dibangunkan dan diberi makan", "Anak babi baru lahir. Akan dikembalikan ke induk", "Anak babi sakit. Akan disuntik"])}),
    ("q15", "繁殖・分娩", 1,
     "母豚が1頭ずつ入る柵（ストール）がならぶ豚舎です。上に青い給餌器（えさの入れもの）があります。",
     {"ja": ("動画の豚舎はどれですか？", ["母豚が1頭ずつ柵に入っている豚舎（ストール）", "子豚がたくさんいる離乳舎", "出荷前の大きい豚の豚舎", "えさを作る部屋"]),
      "en": ("Which kind of pig house is shown in the video?", ["A house where sows are kept one per stall", "A nursery with many piglets", "A house for large pigs before shipping", "A feed preparation room"]),
      "vi": ("Chuồng trong video là loại nào?", ["Chuồng heo nái nhốt từng con trong ô (chuồng cũi)", "Chuồng cai sữa có nhiều heo con", "Chuồng heo lớn trước khi xuất bán", "Phòng trộn thức ăn"]),
      "id": ("Kandang jenis apa yang terlihat dalam video?", ["Kandang induk, satu ekor per sekat (stall)", "Kandang sapih dengan banyak anak babi", "Kandang babi besar sebelum dijual", "Ruang pembuatan pakan"])}),
    ("q16", "子豚・育成", 1,
     "小さい子豚がたくさん入っている豚舎です。丸い給餌器があり、母豚はいません。母豚からはなれた（離乳した）子豚の豚舎です。",
     {"ja": ("動画の豚舎にいる豚はどれですか？", ["母豚からはなれた子豚（離乳子豚）", "妊娠している母豚", "出荷する大きい豚", "生まれたばかりの子豚と母豚"]),
      "en": ("Which pigs are in the house shown in the video?", ["Weaned piglets (separated from the sow)", "Pregnant sows", "Large pigs ready for shipping", "Newborn piglets with their sow"]),
      "vi": ("Heo trong chuồng ở video là loại nào?", ["Heo con đã cai sữa (tách khỏi heo mẹ)", "Heo nái đang mang thai", "Heo lớn sắp xuất bán", "Heo con mới sinh cùng heo mẹ"]),
      "id": ("Babi apa yang ada di kandang dalam video?", ["Anak babi sapih (dipisah dari induk)", "Induk babi bunting", "Babi besar siap jual", "Anak babi baru lahir bersama induk"])}),
    ("q17", "飼養環境・施設", 1,
     "豚舎の外に立っている大きなタンクは飼料タンク（サイロ）です。トラックで運ばれたえさをためて、豚舎の中に送ります。",
     {"ja": ("動画にうつっている大きなタンクは何ですか？", ["えさ（飼料）をためるタンク", "水をためるタンク", "ふん尿をためるタンク", "ガスをためるタンク"]),
      "en": ("What is the large tank shown in the video?", ["A tank that stores feed", "A tank that stores water", "A tank that stores manure", "A tank that stores gas"]),
      "vi": ("Bồn lớn trong video là gì?", ["Bồn chứa thức ăn", "Bồn chứa nước", "Bồn chứa phân", "Bồn chứa gas"]),
      "id": ("Tangki besar dalam video itu apa?", ["Tangki penyimpan pakan", "Tangki penyimpan air", "Tangki penyimpan kotoran", "Tangki penyimpan gas"])}),
    ("q18", "衛生・防疫", 1,
     "きれいな場所（事務所）で、薬のびんに針をさして注射器に薬を吸っています。ほこりや汚れの少ない場所で準備します。",
     {"ja": ("動画では、どこで何をしていますか？", ["事務所で、薬のびんから注射器に薬を吸っている", "豚舎で、豚に注射している", "台所で、薬を水にとかしている", "車の中で、薬を数えている"]),
      "en": ("Where is the video taken, and what is being done?", ["In the office, drawing medicine from a bottle into a syringe", "In the pig house, injecting a pig", "In a kitchen, dissolving medicine in water", "In a car, counting medicine"]),
      "vi": ("Video quay ở đâu và đang làm gì?", ["Ở văn phòng, hút thuốc từ lọ vào ống tiêm", "Ở chuồng, tiêm cho heo", "Ở bếp, hòa thuốc vào nước", "Trong xe, đếm thuốc"]),
      "id": ("Di mana video ini diambil dan apa yang dilakukan?", ["Di kantor, menyedot obat dari botol ke alat suntik", "Di kandang, menyuntik babi", "Di dapur, melarutkan obat dalam air", "Di dalam mobil, menghitung obat"])}),
    ("q19", "衛生・防疫", 1,
     "入口で「外用」と「中用」の長靴をはきかえます。汚れた長靴で中に入ると、外の病気を豚舎に持ち込みます。",
     {"ja": ("動画は豚舎の入口です。白い長靴と汚れた長靴を分けて置くのはなぜですか？", ["外の汚れを豚舎の中に持ち込まないため", "長靴をかわかすため", "サイズで分けるため", "見た目をきれいにするため"]),
      "en": ("The video shows the entrance of a pig house. Why are the white boots and the dirty boots kept separately?", ["So that dirt from outside is not carried into the pig house", "To dry the boots", "To sort them by size", "To make the entrance look tidy"]),
      "vi": ("Video là lối vào chuồng heo. Vì sao ủng trắng và ủng bẩn được để riêng?", ["Để không mang bẩn từ bên ngoài vào chuồng", "Để phơi khô ủng", "Để phân theo cỡ", "Để trông gọn gàng"]),
      "id": ("Video menunjukkan pintu masuk kandang. Mengapa sepatu bot putih dan sepatu bot kotor dipisahkan?", ["Agar kotoran dari luar tidak terbawa ke dalam kandang", "Untuk mengeringkan sepatu bot", "Untuk memisahkan menurut ukuran", "Agar terlihat rapi"])}),
    ("q20", "衛生・防疫", 1,
     "防護服と長靴を着けた人が通路を歩いて、豚のようすを見て回っています（見回り）。元気がない豚や具合の悪い豚がいないか確認します。",
     {"ja": ("動画の人は何をしていますか？", ["防護服を着て、豚舎の中を歩いて豚を見て回っている", "豚舎をそうじしている", "豚を追い出している", "えさを配っている"]),
      "en": ("What is the person in the video doing?", ["Wearing a coverall and walking through the house to check the pigs", "Cleaning the pig house", "Driving the pigs out", "Giving feed"]),
      "vi": ("Người trong video đang làm gì?", ["Mặc đồ bảo hộ, đi trong chuồng để quan sát heo", "Dọn dẹp chuồng", "Lùa heo ra ngoài", "Phát thức ăn"]),
      "id": ("Apa yang dilakukan orang dalam video?", ["Memakai baju pelindung dan berjalan di kandang untuk memeriksa babi", "Membersihkan kandang", "Mengusir babi keluar", "Membagikan pakan"])}),
    # ---- 第2版で追加(q21〜q30)。q27〜q30 は既存の動画に別の角度の問いを付けたもの ----
    ("q21", "繁殖・分娩", 1,
     "分娩柵（母豚が入る柵）のまわりに小さい子豚がいて、子豚用の給餌皿があります。母豚と生まれた子豚がいる分娩舎です。",
     {"ja": ("動画の豚舎はどれですか？", ["母豚と生まれた子豚がいる分娩舎", "子豚だけの離乳舎", "出荷前の大きい豚の豚舎", "えさをためる倉庫"]),
      "en": ("Which kind of pig house is shown in the video?", ["A farrowing house with sows and their newborn piglets", "A nursery with piglets only", "A house for large pigs before shipping", "A feed storage room"]),
      "vi": ("Chuồng trong video là loại nào?", ["Chuồng đẻ có heo mẹ và heo con mới sinh", "Chuồng cai sữa chỉ có heo con", "Chuồng heo lớn trước khi xuất bán", "Kho chứa thức ăn"]),
      "id": ("Kandang jenis apa yang terlihat dalam video?", ["Kandang beranak dengan induk dan anak babi baru lahir", "Kandang sapih hanya anak babi", "Kandang babi besar sebelum dijual", "Gudang pakan"])}),
    ("q22", "肉豚・出荷・肉", 1,
     "1つの豚房に大きい豚が何頭もまとめて入っています（群飼）。母豚を1頭ずつ入れるストールとはちがいます。",
     {"ja": ("動画の豚は、どのように飼われていますか？", ["1つの豚房に何頭もまとめて入っている", "1頭ずつ柵（ストール）に入っている", "母豚と子豚がいっしょに入っている", "1頭ずつ外につながれている"]),
      "en": ("How are the pigs in the video kept?", ["Several pigs together in one pen", "One pig per stall", "A sow together with her piglets", "Each pig tied outside"]),
      "vi": ("Heo trong video được nuôi như thế nào?", ["Nhiều con chung một ô chuồng", "Mỗi con một ô (chuồng cũi)", "Heo mẹ ở cùng heo con", "Mỗi con buộc riêng ngoài trời"]),
      "id": ("Bagaimana babi dalam video dipelihara?", ["Beberapa ekor bersama dalam satu kandang", "Satu ekor per sekat (stall)", "Induk bersama anak-anaknya", "Tiap ekor diikat di luar"])}),
    ("q23", "衛生・防疫", 1,
     "使い終わった注射器を道具のケースに戻しています。床や柵の上に置きっぱなしにすると、汚れたり、なくしたり、豚がかんだりします。",
     {"ja": ("動画では、注射のあとに何をしていますか？", ["注射器を道具のケースにしまっている", "注射器を床に置いている", "注射器を豚房に投げている", "注射器を水につけている"]),
      "en": ("What is being done after the injection in the video?", ["Putting the syringe back into the tool case", "Leaving the syringe on the floor", "Throwing the syringe into the pen", "Putting the syringe in water"]),
      "vi": ("Sau khi tiêm, trong video làm gì?", ["Cất ống tiêm vào hộp dụng cụ", "Để ống tiêm trên sàn", "Ném ống tiêm vào ô chuồng", "Ngâm ống tiêm vào nước"]),
      "id": ("Apa yang dilakukan setelah menyuntik dalam video?", ["Menyimpan alat suntik ke kotak peralatan", "Meletakkan alat suntik di lantai", "Melempar alat suntik ke kandang", "Merendam alat suntik dalam air"])}),
    ("q24", "飼養環境・施設", 1,
     "通路の床がぬれています。ぬれた床はすべりやすく、転んでけがをします。急がず、足もとを見て歩きます。",
     {"ja": ("動画の通路で、歩くときに気をつけることは何ですか？", ["床がぬれていてすべりやすいので、足もとを見てゆっくり歩く", "暗いので走って通る", "豚がいないので何も気をつけなくてよい", "柵の上を歩く"]),
      "en": ("What should you be careful about when walking in the aisle shown?", ["The floor is wet and slippery, so watch your step and walk slowly", "It is dark, so run through", "There are no pigs, so nothing to worry about", "Walk on top of the rails"]),
      "vi": ("Khi đi trong lối đi ở video cần chú ý gì?", ["Sàn ướt, dễ trượt nên nhìn chân và đi chậm", "Tối nên chạy qua", "Không có heo nên không cần chú ý", "Đi trên thanh chắn"]),
      "id": ("Apa yang harus diperhatikan saat berjalan di lorong dalam video?", ["Lantai basah dan licin, jadi perhatikan langkah dan berjalan pelan", "Gelap, jadi lari saja", "Tidak ada babi, jadi tidak perlu hati-hati", "Berjalan di atas pagar"])}),
    ("q25", "衛生・防疫", 1,
     "外から来た人は、豚舎に入る前、農場の敷地を歩くときから防護服と長靴を着けています。豚舎の前で着るのではなく、農場に入るときに着ます。",
     {"ja": ("動画の人は、いつから防護服を着ていますか？", ["豚舎に入る前、農場の道を歩くときから", "豚舎の中に入ってから", "豚にさわる直前だけ", "着ていない"]),
      "en": ("Since when has the person in the video been wearing the coverall?", ["Before entering the pig house, already while walking on the farm road", "Only after entering the pig house", "Only just before touching a pig", "Not wearing one"]),
      "vi": ("Người trong video mặc đồ bảo hộ từ khi nào?", ["Từ trước khi vào chuồng, ngay khi đi trên đường trong trại", "Chỉ sau khi vào trong chuồng", "Chỉ ngay trước khi chạm vào heo", "Không mặc"]),
      "id": ("Sejak kapan orang dalam video memakai baju pelindung?", ["Sebelum masuk kandang, sudah sejak berjalan di jalan peternakan", "Hanya setelah masuk kandang", "Hanya sesaat sebelum menyentuh babi", "Tidak memakai"])}),
    ("q26", "衛生・防疫", 1,
     "水道と桶、かごがある洗い場です。使った道具をここで洗います。",
     {"ja": ("動画の場所は何をする所ですか？", ["使った道具を洗う所", "豚にえさをやる所", "薬をしまう所", "豚を運ぶ所"]),
      "en": ("What is the place in the video used for?", ["Washing used tools", "Feeding pigs", "Storing medicine", "Loading pigs"]),
      "vi": ("Nơi trong video dùng để làm gì?", ["Rửa dụng cụ đã dùng", "Cho heo ăn", "Cất thuốc", "Chuyển heo"]),
      "id": ("Tempat dalam video digunakan untuk apa?", ["Mencuci peralatan bekas pakai", "Memberi makan babi", "Menyimpan obat", "Memuat babi"])}),
    ("q27", "子豚・育成", 1,
     "赤いスプレーは子豚の背中にかけています。背中は上から見えるので、群れの中でも印がすぐ分かります。",
     {"ja": ("動画で、赤いスプレーは子豚のどこにかけていますか？", ["背中", "足", "しっぽ", "おなか"]),
      "en": ("Where on the piglet is the red spray applied in the video?", ["On the back", "On the legs", "On the tail", "On the belly"]),
      "vi": ("Trong video, sơn đỏ được xịt vào đâu trên heo con?", ["Lưng", "Chân", "Đuôi", "Bụng"]),
      "id": ("Di bagian mana anak babi disemprot merah dalam video?", ["Punggung", "Kaki", "Ekor", "Perut"])}),
    ("q28", "肉豚・出荷・肉", 1,
     "動画の豚は大きく、手や体でおさえることができません。だから鼻にワイヤーをかける道具（鼻保定器）で動かないようにします。小さい子豚は手でおさえます。",
     {"ja": ("動画のように鼻にワイヤーをかけて保定するのは、どんな豚ですか？", ["手ではおさえられない大きい豚", "生まれたばかりの子豚", "死んだ豚", "寝ている豚"]),
      "en": ("Which pigs are restrained with a wire snare on the snout, as in the video?", ["Large pigs that cannot be held by hand", "Newborn piglets", "Dead pigs", "Sleeping pigs"]),
      "vi": ("Loại heo nào được cố định bằng dây thòng lọng ở mũi như trong video?", ["Heo lớn không thể giữ bằng tay", "Heo con mới sinh", "Heo đã chết", "Heo đang ngủ"]),
      "id": ("Babi mana yang ditahan dengan jerat kawat di hidung seperti dalam video?", ["Babi besar yang tidak bisa dipegang dengan tangan", "Anak babi baru lahir", "Babi mati", "Babi yang tidur"])}),
    ("q29", "衛生・防疫", 1,
     "長靴の裏（底）を見て、汚れがないか確認しています。長靴の底のみぞには汚れがたまりやすく、病気のもとを運びます。",
     {"ja": ("動画の人は、長靴の裏を見て何を確認していますか？", ["汚れが残っていないか", "サイズが合っているか", "色がきれいか", "値段"]),
      "en": ("What is the person checking by looking at the sole of the boot?", ["Whether any dirt is left", "Whether the size fits", "Whether the color is nice", "The price"]),
      "vi": ("Người trong video nhìn đế ủng để kiểm tra gì?", ["Còn dính bẩn hay không", "Cỡ có vừa không", "Màu có đẹp không", "Giá tiền"]),
      "id": ("Apa yang diperiksa orang itu dengan melihat sol sepatu bot?", ["Apakah masih ada kotoran", "Apakah ukurannya pas", "Apakah warnanya bagus", "Harganya"])}),
    ("q30", "衛生・防疫", 1,
     "とった血は病気の検査に使います。豚が病気にかかっていないか、ワクチンが効いているかを調べます。",
     {"ja": ("動画でとった豚の血は、何に使いますか？", ["病気の検査", "豚のえさ", "肥料", "捨てる"]),
      "en": ("What is the blood taken from the pig in the video used for?", ["Disease testing", "Pig feed", "Fertilizer", "It is thrown away"]),
      "vi": ("Máu lấy từ heo trong video dùng để làm gì?", ["Xét nghiệm bệnh", "Làm thức ăn cho heo", "Làm phân bón", "Vứt bỏ"]),
      "id": ("Darah yang diambil dari babi dalam video digunakan untuk apa?", ["Pemeriksaan penyakit", "Pakan babi", "Pupuk", "Dibuang"])}),
]
# q27〜q30 は既存クリップを使う（id と動画名の対応）
VIDEO_OF = {"q27": "q06", "q28": "q12", "q29": "q01", "q30": "q13"}

DEMO = [
    ("demo01", "子豚・育成", 1,
     "青いスプレーの印は「処置がすんだ」しるしです。印を見て、まだの豚とすんだ豚を見分けます。",
     {"ja": ("動画で子豚の頭に青いスプレーをかけているのはなぜですか？", ["処置がすんだ印をつけるため", "体をひやすため", "虫をよけるため", "きれいにするため"]),
      "en": ("Why is blue spray put on the piglet's head in the video?", ["To mark that the treatment is done", "To cool the body", "To keep insects away", "To clean it"]),
      "vi": ("Vì sao xịt sơn xanh lên đầu heo con trong video?", ["Để đánh dấu đã xử lý xong", "Để làm mát cơ thể", "Để đuổi côn trùng", "Để làm sạch"]),
      "id": ("Mengapa kepala anak babi disemprot biru dalam video?", ["Untuk menandai bahwa tindakan sudah selesai", "Untuk mendinginkan badan", "Untuk mengusir serangga", "Untuk membersihkan"])}),
    ("demo02", "肉豚・出荷・肉", 1,
     "鼻にワイヤーをかけて豚を動かないようにし（鼻保定）、首から注射器で血をとっています（採血）。",
     {"ja": ("動画では豚に何をしていますか？", ["鼻をワイヤーでおさえて（保定して）、首から血をとっている", "鼻に薬をぬっている", "口の中を見ている", "歯を切っている"]),
      "en": ("What is being done to the pig in the video?", ["Holding the snout with a wire snare and taking blood from the neck", "Putting medicine on the snout", "Looking inside the mouth", "Clipping the teeth"]),
      "vi": ("Trong video đang làm gì với con heo?", ["Giữ mũi bằng dây thòng lọng (cố định) và lấy máu ở cổ", "Bôi thuốc lên mũi", "Xem bên trong miệng", "Cắt răng"]),
      "id": ("Apa yang dilakukan pada babi dalam video?", ["Menahan hidung dengan jerat kawat dan mengambil darah dari leher", "Mengoles obat di hidung", "Melihat bagian dalam mulut", "Memotong gigi"])}),
    ("demo03", "衛生・防疫", 1,
     "入口で豚舎用の長靴にはきかえてから中に入っています。外の汚れを持ち込まないためです。",
     {"ja": ("動画の人は豚舎に入る前に何をしましたか？", ["豚舎用の長靴にはきかえた", "手を消毒した", "マスクをつけた", "帽子をかぶった"]),
      "en": ("What did the person do before entering the pig house?", ["Changed into boots for the pig house", "Disinfected the hands", "Put on a mask", "Put on a hat"]),
      "vi": ("Người trong video đã làm gì trước khi vào chuồng heo?", ["Thay sang ủng dành cho chuồng", "Khử trùng tay", "Đeo khẩu trang", "Đội mũ"]),
      "id": ("Apa yang dilakukan orang itu sebelum masuk kandang?", ["Mengganti sepatu bot khusus kandang", "Mendisinfeksi tangan", "Memakai masker", "Memakai topi"])}),
]


def row(item):
    qid, cat, ans, expl, langs = item
    ja_q, ja_o = langs["ja"]
    assert len(ja_o) == 4
    r = [qid, f"videos/{VIDEO_OF.get(qid, qid)}.mp4", cat, 0, 0, ja_q, *ja_o, ans, expl]   # placeholder=0: 2026-09-02 社長判断で全問採用
    for lang in ("en", "vi", "id", "es"):
        if lang in langs:
            q, o = langs[lang]
            assert len(o) == 4, (qid, lang)
            r += [q, *o]
        else:
            r += [""] * 5
    return r


wb = Workbook()
ws = wb.active
ws.title = "本試験"
ws.append(HEAD)
for it in Q:
    ws.append(row(it))
ws2 = wb.create_sheet("デモ")
ws2.append(HEAD)
for it in DEMO:
    ws2.append(row(it))
for sheet in (ws, ws2):
    for c in sheet[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDE8F5")
    sheet.freeze_panes = "C2"
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["F"].width = 50
    sheet.column_dimensions["L"].width = 60
wb.save(XLSX)
print(f"作成: {XLSX}  本試験{len(Q)}問 デモ{len(DEMO)}問")
