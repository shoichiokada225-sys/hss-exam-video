#!/usr/bin/env python3
"""tools/questions.xlsx → data/questions.json / data/demo-questions.json を生成する。

使い方:  python tools/build_from_xlsx.py            # 変換して書き出す
         python tools/build_from_xlsx.py --check    # 書き出さず検査だけ
         python tools/build_from_xlsx.py --template # 空のテンプレート xlsx を作る(既存があれば上書きしない)

シート「本試験」「デモ」の列(1行目が見出し):
  id | video | category | required | placeholder |
  question_ja | opt1_ja | opt2_ja | opt3_ja | opt4_ja | answer | explanation_ja |
  question_en | opt1_en .. opt4_en | question_vi | opt1_vi .. | question_id | opt1_id .. | question_es | opt1_es ..
  - answer は 1〜4 の番号(日本語選択肢の何番目が正解か)。全言語の選択肢はこの並びに揃える(順序を崩すと採点事故)
  - required=1 の問題は毎回必ず出題(問題数より多いプールから抽出する場合のみ意味を持つ)
  - placeholder=1 は仮問題または未監修の問題。validate.py --release が本番前に検出して止める（社長・獣医の確認後に0にする）
  - 翻訳セルが空なら、その言語は日本語で表示される(アプリ側フォールバック)
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, ValueError):
    pass

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "tools" / "questions.xlsx"
LANGS = [("en", "en"), ("vi", "vi"), ("id", "ind"), ("es", "es")]  # (列サフィックス, JSONキー)
CATEGORIES = ["繁殖・分娩", "子豚・育成", "衛生・防疫", "飼料・栄養", "飼養環境・施設", "肉豚・出荷・肉", "経営理念・行動規範"]

HEAD = ["id", "video", "category", "required", "placeholder",
        "question_ja", "opt1_ja", "opt2_ja", "opt3_ja", "opt4_ja", "answer", "explanation_ja"]
for _suf, _ in LANGS:
    HEAD += [f"question_{_suf}"] + [f"opt{i}_{_suf}" for i in range(1, 5)]

TRUTHY = ("1", "true", "TRUE", "○", "1.0")


def cell(v):
    return "" if v is None else str(v).strip()


def rows_to_questions(ws, sheet_label):
    head = [cell(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(head)}
    missing = [h for h in HEAD if h not in idx]
    if missing:
        raise SystemExit(f"[{sheet_label}] 見出しが足りません: {missing}")
    out = []
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def g(h):
            return cell(row[idx[h]]) if idx[h] < len(row) else ""
        if not g("question_ja") and not g("video"):
            continue  # 空行
        if "review_status" in idx and g("review_status") == "del":
            print(f"  [{sheet_label}] {r}行目 {g('id')}: review_status=del のため除外")
            continue
        opts = [g(f"opt{i}_ja") for i in range(1, 5)]
        opts = [o for o in opts if o]
        try:
            ans = int(float(g("answer")))
        except ValueError:
            raise SystemExit(f"[{sheet_label}] {r}行目: answer が数字ではありません: {g('answer')!r}")
        if not (1 <= ans <= len(opts)):
            raise SystemExit(f"[{sheet_label}] {r}行目: answer={ans} が選択肢数{len(opts)}の範囲外")
        q = {
            "id": g("id") or f"{sheet_label}-{r - 1:02d}",
            "video": g("video"),
            "category": g("category"),
            "question": g("question_ja"),
            "options": opts,
            "answer": opts[ans - 1],
            "explanation": g("explanation_ja"),
        }
        if g("required") in TRUTHY:
            q["required"] = True
        if g("placeholder") in TRUTHY:
            q["placeholder"] = True
        for suf, key in LANGS:
            tq = g(f"question_{suf}")
            topts = [g(f"opt{i}_{suf}") for i in range(1, len(opts) + 1)]
            if tq and all(topts):
                q[key] = {"question": tq, "options": topts}
            elif tq or any(topts):
                raise SystemExit(f"[{sheet_label}] {r}行目: {suf} の翻訳が中途半端です(問題文と選択肢{len(opts)}個を全部埋めるか、全部空にする)")
        out.append(q)
    return out


def build(check_only=False):
    if not XLSX.exists():
        raise SystemExit(f"{XLSX} がありません。--template で作れます")
    wb = load_workbook(XLSX, data_only=True)
    targets = {"本試験": ROOT / "data" / "questions.json", "デモ": ROOT / "data" / "demo-questions.json"}
    for sheet, dest in targets.items():
        if sheet not in wb.sheetnames:
            print(f"シート「{sheet}」なし: スキップ")
            continue
        qs = rows_to_questions(wb[sheet], sheet)
        ph = sum(1 for q in qs if q.get("placeholder"))
        print(f"{sheet}: {len(qs)}問 (仮問題 {ph})")
        if not check_only:
            dest.write_text(json.dumps(qs, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
            print(f"  -> {dest.relative_to(ROOT)}")


def template():
    if XLSX.exists():
        raise SystemExit(f"{XLSX} は既にあります(上書きしません)")
    wb = Workbook()
    ws = wb.active
    ws.title = "本試験"
    ws.append(HEAD)
    ws2 = wb.create_sheet("デモ")
    ws2.append(HEAD)
    wb.save(XLSX)
    print(f"作成: {XLSX}")


if __name__ == "__main__":
    if "--template" in sys.argv:
        template()
    else:
        build(check_only="--check" in sys.argv)
