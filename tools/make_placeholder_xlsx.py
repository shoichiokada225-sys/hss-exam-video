#!/usr/bin/env python3
"""骨組み用の仮問題(本試験30問+デモ3問)を tools/questions.xlsx に書く。
本物の問題ができたらこのスクリプトは不要(Excelを直接編集→build_from_xlsx.py)。
既存の questions.xlsx がある場合は --force を付けない限り上書きしない。
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_from_xlsx import CATEGORIES, HEAD, XLSX  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, ValueError):
    pass

if XLSX.exists() and "--force" not in sys.argv:
    raise SystemExit(f"{XLSX} は既にあります(--force で上書き)")

# 仮の問題文。観察型の出題テンプレートを4種類ローテーションし、本物に差し替える時の型にする
TEMPLATES = [
    ("動画の作業で、正しくないのはどれですか？",
     ["【仮】誤っている手順A", "【仮】正しい手順B", "【仮】正しい手順C", "【仮】正しい手順D"], 1),
    ("動画の豚に見られる異常はどれですか？",
     ["【仮】異常なし", "【仮】異常の候補B", "【仮】異常の候補C", "【仮】異常の候補D"], 2),
    ("動画の場面で、次にすべき行動はどれですか？",
     ["【仮】行動A", "【仮】行動B", "【仮】行動C", "【仮】行動D"], 3),
    ("動画の場面で、危険なのはどれですか？",
     ["【仮】危険A", "【仮】危険B", "【仮】危険C", "【仮】危険D"], 4),
]
EN = [
    ("Which action in the video is NOT correct?",
     ["[TBD] wrong step A", "[TBD] correct step B", "[TBD] correct step C", "[TBD] correct step D"]),
    ("What abnormality is seen in the pig in the video?",
     ["[TBD] nothing abnormal", "[TBD] candidate B", "[TBD] candidate C", "[TBD] candidate D"]),
    ("In the situation shown, what should be done next?",
     ["[TBD] action A", "[TBD] action B", "[TBD] action C", "[TBD] action D"]),
    ("What is dangerous in the situation shown?",
     ["[TBD] hazard A", "[TBD] hazard B", "[TBD] hazard C", "[TBD] hazard D"]),
]


def row(i, prefix):
    t = TEMPLATES[i % 4]
    en = EN[i % 4]
    vid = f"videos/dummy-{prefix}{i + 1:02d}.mp4"
    r = [f"{prefix}{i + 1:02d}", vid, CATEGORIES[i % len(CATEGORIES)], 0, 1,
         f"【仮問題{i + 1}】" + t[0], *t[1], t[2], f"【仮】解説{i + 1}。本物の動画ができたらここに根拠を書く。"]
    # en だけ仮訳を入れて多言語の通り道を確認できるようにする。vi/id/es は空=日本語フォールバック
    r += [f"[Placeholder {i + 1}] " + en[0], *en[1]]
    r += [""] * 5 * 3
    return r


wb = Workbook()
ws = wb.active
ws.title = "本試験"
ws.append(HEAD)
for i in range(30):
    ws.append(row(i, "q"))
ws2 = wb.create_sheet("デモ")
ws2.append(HEAD)
for i in range(3):
    ws2.append(row(i, "demo"))
for sheet in (ws, ws2):
    for c in sheet[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDE8F5")
    sheet.freeze_panes = "C2"
    sheet.column_dimensions["B"].width = 26
    sheet.column_dimensions["F"].width = 48
XLSX.parent.mkdir(exist_ok=True)
wb.save(XLSX)
print(f"作成: {XLSX}")
