#!/usr/bin/env python3
"""review.html でダウンロードした判定JSONを tools/questions.xlsx に反映する。

使い方:  python tools/apply_review.py ~/Downloads/review-2026-09-10.json
         python tools/apply_review.py review.json --dry-run   # 変更内容の表示だけ

反映内容:
  - status=ok  → placeholder 列を 0 にする（監修済み）
  - status=fix → placeholder は 1 のまま。列「review_memo」にメモを書く（人が Excel で直す）
  - status=del → 行は消さず、列「review_status」に del と書く（build 時に除外される）
  - 列 review_status / review_memo / review_date が無ければ末尾に追加する
反映後は python tools/build_from_xlsx.py を実行すること。
"""
import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_from_xlsx import XLSX  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, ValueError):
    pass

EXTRA = ["review_status", "review_memo", "review_date"]


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    dry = "--dry-run" in sys.argv
    if not args:
        raise SystemExit(__doc__)
    review = json.loads(Path(args[0]).read_text(encoding="utf-8"))
    items = {it["id"]: it for it in review.get("items", []) if it.get("status")}
    if not items:
        raise SystemExit("判定が1件もありません")
    wb = load_workbook(XLSX)
    today = date.today().isoformat()
    changed = 0
    for ws in wb.worksheets:
        head = [str(c.value or "").strip() for c in ws[1]]
        if "id" not in head or "placeholder" not in head:
            continue
        for col in EXTRA:
            if col not in head:
                ws.cell(row=1, column=len(head) + 1, value=col)
                head.append(col)
        ci = {h: i + 1 for i, h in enumerate(head)}
        for r in range(2, ws.max_row + 1):
            qid = str(ws.cell(row=r, column=ci["id"]).value or "").strip()
            it = items.get(qid)
            if not it:
                continue
            st = it["status"]
            memo = it.get("memo", "")
            before = ws.cell(row=r, column=ci["placeholder"]).value
            if st == "ok":
                ws.cell(row=r, column=ci["placeholder"], value=0)
            ws.cell(row=r, column=ci["review_status"], value=st)
            ws.cell(row=r, column=ci["review_memo"], value=memo)
            ws.cell(row=r, column=ci["review_date"], value=today)
            changed += 1
            print(f"[{ws.title}] {qid}: {st}  placeholder {before}->{ws.cell(row=r, column=ci['placeholder']).value}  {memo[:40]}")
    if dry:
        print(f"(dry-run) {changed}件。書き込みなし")
        return
    wb.save(XLSX)
    print(f"{changed}件を {XLSX.name} に反映。次: python tools/build_from_xlsx.py")


if __name__ == "__main__":
    main()
